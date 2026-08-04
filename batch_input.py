"""Batch input parsing for the sequential MFDS review app."""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any, Iterable

MAX_BATCH_SIZE = 50

FIELD_ALIASES = {
    "record_id": ("record_id", "id", "레코드 ID", "레코드ID", "식별자"),
    "title": ("title", "제목", "게시물 제목"),
    "body_text": ("body_text", "body", "내용", "본문", "게시물 내용"),
    "source_url": (
        "source_url",
        "url",
        "원문 URL",
        "원문URL",
        "사이트주소",
        "URL",
    ),
    "platform": ("platform", "사이트명", "플랫폼"),
    "product_name": ("product_name", "제품명", "추가요청 (제품명)"),
}


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _header_key(value: Any) -> str:
    return "".join(_clean(value).lower().split())


def _field_map(headers: Iterable[Any]) -> dict[str, int]:
    normalized = [_header_key(header) for header in headers]
    aliases = {
        _header_key(alias): field
        for field, values in FIELD_ALIASES.items()
        for alias in values
    }
    return {
        field: index
        for index, header in enumerate(normalized)
        if (field := aliases.get(header)) is not None
    }


def _normalize_row(row: dict[str, Any], index: int) -> dict[str, Any]:
    normalized_row = {_header_key(key): value for key, value in row.items()}
    values: dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            candidate = normalized_row.get(_header_key(alias))
            if _clean(candidate):
                values[field] = _clean(candidate)
                break
        values.setdefault(field, "")

    values["record_id"] = values["record_id"] or f"BATCH-{index:03d}"
    if not values["title"] and not values["body_text"] and not values["source_url"]:
        raise ValueError(f"{index}행에 제목·본문·원문 URL 중 하나가 필요합니다.")
    return values


def _rows_from_csv(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV 헤더 행을 찾을 수 없습니다.")
    return [dict(row) for row in reader]


def _rows_from_jsonl(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig")
    rows: list[dict[str, Any]] = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        if not line.strip():
            continue
        try:
            value = json.loads(line)
        except json.JSONDecodeError as error:
            raise ValueError(f"JSONL {line_number}행 형식 오류: {error.msg}") from error
        if not isinstance(value, dict):
            raise ValueError(f"JSONL {line_number}행은 JSON 객체여야 합니다.")
        rows.append(value)
    return rows


def _rows_from_xlsx(raw: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError("XLSX 입력에는 openpyxl 설치가 필요합니다.") from error

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration as error:
        raise ValueError("XLSX에 데이터가 없습니다.") from error
    if not any(_clean(value) for value in headers):
        raise ValueError("XLSX 첫 행에서 헤더를 찾을 수 없습니다.")
    field_map = _field_map(headers)
    rows: list[dict[str, Any]] = []
    for row in iterator:
        if not any(_clean(value) for value in row):
            continue
        rows.append(
            {
                field: row[column_index] if column_index < len(row) else ""
                for field, column_index in field_map.items()
            }
        )
    return rows


def parse_batch_bytes(filename: str, raw: bytes) -> list[dict[str, Any]]:
    """Parse CSV, JSONL, TXT, or XLSX rows into pipeline source objects."""

    suffix = Path(filename).suffix.lower()
    if suffix in {".jsonl", ".ndjson"}:
        rows = _rows_from_jsonl(raw)
    elif suffix == ".xlsx":
        rows = _rows_from_xlsx(raw)
    else:
        rows = _rows_from_csv(raw)
    return normalize_batch_rows(rows)


def parse_batch_text(text: str) -> list[dict[str, Any]]:
    """Parse pasted JSONL or CSV text."""

    raw = text.encode("utf-8")
    stripped = text.lstrip()
    if stripped.startswith("{"):
        rows = _rows_from_jsonl(raw)
    else:
        rows = _rows_from_csv(raw)
    return normalize_batch_rows(rows)


def normalize_batch_rows(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = [_normalize_row(row, index) for index, row in enumerate(rows, start=1)]
    if not normalized:
        raise ValueError("처리할 배치 행이 없습니다.")
    if len(normalized) > MAX_BATCH_SIZE:
        raise ValueError(f"한 번에 최대 {MAX_BATCH_SIZE}건까지 처리할 수 있습니다.")
    return normalized
